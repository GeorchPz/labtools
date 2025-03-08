# -*- coding: utf-8 -*-
"""
Created on Mon Jun 22 21:02:09 2020

@author: Jorge Pottiez
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def xl_enhancer(path):
    '''Tweaks cells' size & alignment'''
    wb = load_workbook(path)     
    # Worksheet
    all_ws_names = wb.sheetnames

    for name in all_ws_names:
        ws = wb[name]
        
        column_widths = []
        for column in ws.columns:
            # Horizontal text alignment
            for cell in column:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column width size
            length = max( len( str(cell.value) ) for cell in column)
            column_widths.append(length)
        
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width
    
    wb.save(path)