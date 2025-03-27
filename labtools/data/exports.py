# -*- coding: utf-8 -*-
"""
Created on Mon Jun 22 21:02:09 2020

@author: Jorge Pottiez
"""

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from .path_utils import get_file_path

def unload_excel(filename, dataframes, sheet_names=None, directory='data', enhance=True):
    '''
    Write multiple DataFrames to an Excel file with formatting enhancement.
    
    Parameters:
        - dataframes : list of pandas.DataFrame, data to write
        - filename : str, Excel filename (without .xlsx extension)
        - sheet_names : list of str, optional, names of the sheets (default: None, will use Sheet1, Sheet2, etc.)
        - directory : str, optional, directory path (default: module's RESULTS_DIR)
    Returns:
        - str, full file path of the written Excel file
    '''
    if isinstance(dataframes, pd.DataFrame):
        dataframes = [dataframes]
    
    if sheet_names is None:
        n_dfs = len(dataframes)
        sheet_names = [f'Sheet{i+1}' for i in range(n_dfs)]

    file_path = get_file_path(filename, extension='.xlsx', directory=directory)
    # file_path = get_file_path(filename+ '_©', extension='.xlsx', directory=directory)

    # Use context manager to handle closing automatically
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        for dataframe, sheet_name in zip(dataframes, sheet_names):
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    
    if enhance:
        _enhance_excel(file_path)
        
    return file_path

def unload_csv(filename, dataframes, delimiter=',', directory='results'):
    '''
    Write multiple DataFrames to a CSV file.

    '''

    file_path = get_file_path(filename, extension='.csv', directory=directory)

    # Check if all dataframes have the same headers
    headers = [list(df.columns) for df in dataframes]
    same_headers = all(header == headers[0] for header in headers)

    if same_headers:
        # Append DataFrames one on top of the other
        for i, dataframe in enumerate(dataframes):
            if i == 0:
                dataframe.to_csv(file_path, sep=delimiter, index=False)
            else:
                dataframe.to_csv(file_path, sep=delimiter, mode='a', header=False, index=False)
                # mode='a': append mode
    else:
        # Concatenate DataFrames side by side
        combined_df = pd.concat(dataframes, axis=1)
        
        combined_df.to_csv(file_path, sep=delimiter, index=False)
    
    return file_path

def _enhance_excel(file_path):
    '''
    Enhances the Excel file by:
        - centering the text
        - adjusting the column width to the longest cell width
    
    Parameters
    ----------
    file_path : str
        Path to the Excel file.
    '''
    wb = load_workbook(file_path)     

    for ws in wb.worksheets:
        column_widths = []
        for column in ws.columns:
            # Center align all cells
            for cell in column:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column width size
            length = max( len( str(cell.value) if cell.value is not None else "") for cell in column)
            column_widths.append(max(length, 10)) # Minimum width of 10
        
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width
    
    wb.save(file_path)